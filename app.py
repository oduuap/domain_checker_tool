#!/usr/bin/env python3
"""
DOMAIN FINDER WEB APP
Web interface để tìm expired domains với traffic
"""

from flask import Flask, render_template, request, jsonify, send_file, Response
from flask_cors import CORS
import json
import time
import os
from datetime import datetime
from threading import Thread, Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
import queue
import whois
import dns.resolver
import requests
import re
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from urllib.parse import quote
from keyword_database import KeywordGenerator

app = Flask(__name__)
CORS(app)

# Initialize keyword generator
keyword_gen = KeywordGenerator()

# API Keys
RAPIDAPI_KEY = "3267466f8dmshf9b9f3bb87d2afcp10c10bjsnccecb46bc96a"
C99_API_KEY = "POM2S-E8ZC6-7KFVA-VH8TP"

# Global variables cho progress tracking
search_progress = {
    'status': 'idle',
    'current': 0,
    'total': 0,
    'message': '',
    'domains_found': [],
    'current_domain': ''
}

class DomainChecker:
    def __init__(self, rapidapi_key=None, c99_api_key=None):
        self.rapidapi_key = rapidapi_key
        self.c99_api_key = c99_api_key

    def check_domain_availability(self, domain: str) -> Dict:
        """Kiểm tra domain availability với phân loại CHI TIẾT"""
        result = {
            'domain': domain,
            'available': None,
            'creation_date': None,
            'age_years': None,
            'registrar': None,
            'status': None,
            'status_type': None,      # available/pending_delete/redemption/auction/registered
            'status_badge': None,     # Badge hiển thị
            'status_note': None,      # Note chi tiết
            'action_type': None,      # buy/backorder/auction/contact
            'action_url': None        # URL action
        }

        try:
            w = whois.whois(domain)

            if w.domain_name:
                result['available'] = False
                result['registrar'] = w.registrar

                # Parse WHOIS status codes
                status_codes = []
                if w.status:
                    if isinstance(w.status, list):
                        status_codes = [str(s).lower() for s in w.status]
                    else:
                        status_codes = [str(w.status).lower()]

                status_str = ' '.join(status_codes)

                # PHÂN LOẠI STATUS CHI TIẾT
                if 'pendingdelete' in status_str or 'pending delete' in status_str:
                    # PENDING DELETE - Sắp xóa trong 5 ngày
                    result['status'] = 'Pending Delete'
                    result['status_type'] = 'pending_delete'
                    result['status_badge'] = '⏳ Pending Delete'
                    result['status_note'] = 'Sẽ xóa trong 5 ngày - Có thể đặt backorder'
                    result['action_type'] = 'backorder'
                    result['action_url'] = f'https://www.dropcatch.com/domain/{domain}'

                elif 'redemption' in status_str:
                    # REDEMPTION PERIOD - Chủ cũ có 30 ngày chuộc
                    result['status'] = 'Redemption Period'
                    result['status_type'] = 'redemption'
                    result['status_badge'] = '⚠️ Redemption'
                    result['status_note'] = 'Chủ cũ có 30 ngày chuộc - Có thể đặt backorder'
                    result['action_type'] = 'backorder'
                    result['action_url'] = f'https://www.dropcatch.com/domain/{domain}'

                elif 'auction' in status_str:
                    # AUCTION - Đang đấu giá
                    result['status'] = 'Auction'
                    result['status_type'] = 'auction'
                    result['status_badge'] = '🔨 Auction'
                    result['status_note'] = 'Đang đấu giá - Có thể tham gia'
                    result['action_type'] = 'auction'
                    result['action_url'] = f'https://auctions.godaddy.com/trp/search?q={domain}'

                else:
                    # REGISTERED (OK) - Bình thường, không mua được
                    result['status'] = 'Registered'
                    result['status_type'] = 'registered'
                    result['status_badge'] = '🔒 Registered'
                    result['status_note'] = 'Đã có chủ - Liên hệ để mua'
                    result['action_type'] = 'contact'
                    result['action_url'] = f'https://www.namecheap.com/domains/registration/results/?domain={domain}'

                # Creation date & age
                creation_date = w.creation_date
                if isinstance(creation_date, list):
                    creation_date = creation_date[0]

                if creation_date:
                    result['creation_date'] = creation_date.strftime('%Y-%m-%d') if isinstance(creation_date, datetime) else str(creation_date)
                    if isinstance(creation_date, datetime):
                        creation_date_naive = creation_date.replace(tzinfo=None) if creation_date.tzinfo else creation_date
                        age = datetime.now() - creation_date_naive
                        result['age_years'] = round(age.days / 365.25, 1)
                    else:
                        result['age_years'] = 'N/A'
                else:
                    result['age_years'] = 'N/A'
            else:
                # AVAILABLE - Mua ngay
                result['available'] = True
                result['status'] = 'Available'
                result['status_type'] = 'available'
                result['status_badge'] = '✅ Available'
                result['status_note'] = 'Có thể mua ngay'
                result['action_type'] = 'buy'
                result['action_url'] = f'https://www.namecheap.com/domains/registration/results/?domain={domain}'

        except Exception as e:
            error_msg = str(e).lower()
            if 'no match' in error_msg or 'not found' in error_msg or 'no data' in error_msg:
                # AVAILABLE - Mua ngay
                result['available'] = True
                result['status'] = 'Available'
                result['status_type'] = 'available'
                result['status_badge'] = '✅ Available'
                result['status_note'] = 'Có thể mua ngay'
                result['action_type'] = 'buy'
                result['action_url'] = f'https://www.namecheap.com/domains/registration/results/?domain={domain}'

        return result

    def check_wayback_history(self, domain: str) -> Dict:
        """Kiểm tra lịch sử trên Wayback Machine"""
        result = {
            'snapshot_count': 0,
            'first_archive': None,
            'last_archive': None,
            'age_years': 0,
            'has_history': False
        }

        try:
            url = f"http://web.archive.org/cdx/search/cdx?url={domain}&output=json&limit=10000"
            response = requests.get(url, timeout=15)

            if response.status_code == 200:
                data = response.json()

                if len(data) > 1:
                    result['snapshot_count'] = len(data) - 1
                    result['has_history'] = True

                    first_timestamp = data[1][1]
                    last_timestamp = data[-1][1]

                    result['first_archive'] = first_timestamp[:8]
                    result['last_archive'] = last_timestamp[:8]

                    first_year = int(first_timestamp[:4])
                    current_year = datetime.now().year
                    result['age_years'] = current_year - first_year

        except Exception as e:
            pass

        return result

    def check_seo_metrics_rapidapi(self, domain: str) -> Dict:
        """Kiểm tra DR/UR/SEO metrics qua RapidAPI"""
        result = {
            'domain_rating': 0,
            'url_rating': 0,
            'organic_traffic': 0,
            'backlinks': 0,
            'referring_domains': 0,
            'has_seo_value': False
        }

        if not self.rapidapi_key:
            return result

        try:
            url = "https://seo-api-dr-rd-rank-keywords-backlinks.p.rapidapi.com/url-metrics"

            headers = {
                'x-rapidapi-host': 'seo-api-dr-rd-rank-keywords-backlinks.p.rapidapi.com',
                'x-rapidapi-key': self.rapidapi_key
            }

            params = {'url': f'https://{domain}'}

            response = requests.get(url, headers=headers, params=params, timeout=15)

            if response.status_code == 200:
                data = response.json()

                # Parse ĐÚNG structure của RapidAPI response
                if data.get('success') and 'data' in data:
                    api_data = data['data']

                    # Lấy domain metrics
                    domain_metrics = api_data.get('domain', {})
                    page_metrics = api_data.get('page', {})

                    result['domain_rating'] = domain_metrics.get('domainRating', 0) or 0
                    result['url_rating'] = page_metrics.get('urlRating', 0) or 0
                    # GIỮ TRAFFIC DẠNG FLOAT để không mất traffic nhỏ (< 1)
                    result['organic_traffic'] = float(domain_metrics.get('trafficVol', 0) or 0)
                    result['backlinks'] = domain_metrics.get('backlinks', 0) or 0
                    result['referring_domains'] = domain_metrics.get('refDomains', 0) or 0
                    result['has_seo_value'] = result['domain_rating'] > 0 or result['organic_traffic'] > 0

        except Exception as e:
            pass

        return result

    def check_registrar_availability(self, domain: str) -> Dict:
        """Kiểm tra domain có mua được không tại Namecheap"""
        result = {
            'purchasable': False,
            'price': None,
            'url': None,
            'is_premium': False
        }

        try:
            search_url = f"https://www.namecheap.com/domains/registration/results/?domain={domain}"
            result['url'] = search_url

            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
            }

            response = requests.get(search_url, headers=headers, timeout=15)

            if response.status_code == 200:
                content = response.text.lower()

                if 'available' in content and 'add to cart' in content:
                    result['purchasable'] = True

                    price_match = re.search(r'\$(\d+\.\d+)', response.text)
                    if price_match:
                        result['price'] = f"${price_match.group(1)}"

                if 'premium' in content:
                    result['is_premium'] = True

            time.sleep(1)

        except Exception as e:
            pass

        return result

    def generate_keyword_variations(self, keyword: str, max_variations: int = 20) -> List[str]:
        """Tạo variations từ keyword"""
        variations = [keyword]

        suffixes = ['hub', 'app', 'pro', 'web', 'net', 'zone', 'spot', 'land', 'world']
        prefixes = ['my', 'get', 'the', 'top', 'best', 'new', 'hot']

        for suffix in suffixes:
            variations.append(f"{keyword}{suffix}")

        for prefix in prefixes:
            variations.append(f"{prefix}{keyword}")

        return list(dict.fromkeys(variations))[:max_variations]

    def fetch_subdomains_c99_api(self, domain: str) -> List[str]:
        """
        Lấy subdomains qua C99 API - CHÍNH XÁC & NHANH

        Args:
            domain: TLD (vd: 'sa.com', 'ca.com')

        Returns:
            List subdomains
        """
        subdomains = []

        if not self.c99_api_key:
            return subdomains

        try:
            # C99 Subdomain Finder API
            url = "https://api.c99.nl/subdomainfinder"

            params = {
                'key': self.c99_api_key,
                'domain': domain,
                'json': 'true'
            }

            print(f"  📡 Calling C99 API for {domain}...", end=' ')

            response = requests.get(url, params=params, timeout=30)

            if response.status_code == 200:
                data = response.json()

                # Check response success
                if data.get('success') == True or data.get('success') == 1:
                    # Parse subdomains từ response
                    if 'subdomains' in data:
                        subdomains = data['subdomains']
                    elif 'result' in data:
                        result = data['result']
                        if isinstance(result, list):
                            subdomains = result
                        elif isinstance(result, dict) and 'subdomains' in result:
                            subdomains = result['subdomains']

                    print(f"✓ Found {len(subdomains)} subdomains")

                else:
                    error = data.get('error', 'Unknown error')
                    print(f"✗ API Error: {error}")

            else:
                print(f"✗ HTTP {response.status_code}")

        except Exception as e:
            print(f"✗ Error: {str(e)[:50]}")

        return subdomains

    def fetch_domains_from_c99(self, tld: str) -> List[str]:
        """
        Lấy danh sách domains từ c99.nl API

        Args:
            tld: TLD cần lấy (vd: 'ca.com', 'sa.com', 'ru.com')

        Returns:
            List domains
        """
        domains = []

        # Try C99 API
        if self.c99_api_key:
            print(f"\n📡 Fetching domains from C99 API for {tld}...")
            raw_domains = self.fetch_subdomains_c99_api(tld)

            if raw_domains:
                # Parse domains - handle multiple formats
                parsed_domains = []

                for item in raw_domains:
                    try:
                        if isinstance(item, str):
                            # Format 1: ["domain.com", "domain2.com"]
                            parsed_domains.append(item.lower().strip())
                        elif isinstance(item, dict):
                            # Format 2: [{"domain": "domain.com", "ip": "1.2.3.4"}]
                            if 'domain' in item:
                                parsed_domains.append(item['domain'].lower().strip())
                            elif 'subdomain' in item:
                                parsed_domains.append(item['subdomain'].lower().strip())
                        elif isinstance(item, list) and len(item) > 0:
                            # Format 3: [["domain.com", "1.2.3.4"], ...]
                            parsed_domains.append(str(item[0]).lower().strip())
                    except Exception as e:
                        # Skip invalid items
                        continue

                # Remove duplicates
                domains = list(set(parsed_domains))

                # Filter bỏ domains quá dài (>50 chars) và không hợp lệ
                domains = [d for d in domains if d and len(d) <= 50 and '.' in d]

                print(f"  ✓ Got {len(domains)} valid domains from C99.NL")
                return domains

        print(f"  ✗ Could not fetch domains from C99 API")
        return domains


def search_domains_background(keywords, tlds, max_check, min_dr, search_id, mode='keyword', c99_domains=None):
    """Background task để search domains"""
    global search_progress

    try:
        search_progress['status'] = 'running'
        search_progress['message'] = 'Đang khởi tạo...'

        print(f"\n{'='*70}")
        print(f"🔍 STARTING SEARCH - ID: {search_id}")
        print(f"{'='*70}")
        print(f"Mode: {mode}")
        print(f"Keywords: {keywords}")
        print(f"TLDs: {tlds}")
        print(f"Max check: {max_check}")
        print(f"Min DR: {min_dr}")
        if mode == 'c99' and c99_domains:
            print(f"C99 Domains: {len(c99_domains)} domains")
        print(f"{'='*70}\n")

        checker = DomainChecker(rapidapi_key=RAPIDAPI_KEY, c99_api_key=C99_API_KEY)

        # Generate domains based on mode
        if mode == 'c99' and c99_domains:
            # C99 Mode: Check HẾT domains từ C99.NL - Lấy domains có Traffic/DR/Backlinks
            search_progress['message'] = 'Dùng domains từ C99.NL...'
            domains_to_check = c99_domains  # Check hết tất cả!
            print(f"✓ Will check ALL {len(domains_to_check)} domains from C99.NL")
            print(f"✓ Will show domains with Traffic OR DR OR Backlinks (Traffic prioritized)")

        else:
            # Keyword Mode: Generate từ keywords
            search_progress['message'] = 'Đang tạo danh sách domains...'
            all_keywords = []
            for kw in keywords:
                variations = checker.generate_keyword_variations(kw.strip().lower(), max_variations=15)
                all_keywords.extend(variations)

            all_keywords = list(dict.fromkeys(all_keywords))[:100]

            all_domains = []
            for kw in all_keywords:
                for tld in tlds:
                    all_domains.append(f"{kw}.{tld}")

            domains_to_check = all_domains[:max_check]

        search_progress['total'] = len(domains_to_check)
        search_progress['message'] = f'Sẽ kiểm tra {len(domains_to_check)} domains...'

        # Check từng domain - PARALLEL MODE
        found_domains = []
        found_domains_lock = Lock()  # Thread-safe access
        checked_count = 0
        skipped_registered = 0
        skipped_no_history = 0
        skipped_low_dr = 0
        skipped_not_purchasable = 0

        # Hàm check 1 domain (sẽ chạy song song)
        def check_single_domain_c99(domain, idx):
            """Check 1 domain trong C99 mode - thread-safe"""
            try:
                # Check SEO metrics
                seo_metrics = checker.check_seo_metrics_rapidapi(domain)

                # LẤY NẾU CÓ: Traffic > 0 HOẶC DR > 0 HOẶC Backlinks > 0 ⭐
                has_traffic = seo_metrics['organic_traffic'] > 0
                has_dr = seo_metrics['domain_rating'] > 0
                has_backlinks = seo_metrics['backlinks'] > 0

                if has_traffic or has_dr or has_backlinks:
                    # CHECK WHOIS CHI TIẾT (Available/Pending Delete/Redemption/Auction/Registered)
                    whois_result = checker.check_domain_availability(domain)

                    # Lấy thông tin chi tiết từ WHOIS
                    is_available = whois_result.get('available', False)
                    status_type = whois_result.get('status_type', 'unknown')
                    status_badge = whois_result.get('status_badge', 'N/A')
                    status_note = whois_result.get('status_note', '')
                    action_type = whois_result.get('action_type', 'contact')
                    action_url = whois_result.get('action_url', f"https://www.namecheap.com/domains/registration/results/?domain={domain}")

                    # Set price based on status
                    if status_type == 'available':
                        price_estimate = '$10-15/year'
                    elif status_type in ['pending_delete', 'redemption']:
                        price_estimate = 'Backorder $69+'
                    elif status_type == 'auction':
                        price_estimate = 'Auction - Varies'
                    else:
                        price_estimate = 'Contact Owner'

                    # Return domain data với STATUS CHI TIẾT
                    return {
                        'domain': domain,
                        'domain_rating': seo_metrics['domain_rating'],
                        'url_rating': seo_metrics['url_rating'],
                        'organic_traffic': seo_metrics['organic_traffic'],
                        'backlinks': seo_metrics['backlinks'],
                        'referring_domains': seo_metrics['referring_domains'],
                        'snapshot_count': 0,
                        'first_archive': 'N/A',
                        'age_years': whois_result.get('age_years', 0),
                        'available': is_available,
                        'whois_status': status_badge,
                        'status_type': status_type,
                        'status_note': status_note,
                        'action_type': action_type,
                        'price': price_estimate,
                        'purchase_url': action_url
                    }
            except Exception as e:
                print(f"Error checking {domain}: {e}")
                return None

        # CHECK SONG SONG với ThreadPoolExecutor
        if mode == 'c99':
            print(f"\n{'='*70}")
            print(f"🚀 PARALLEL MODE: Checking {len(domains_to_check)} domains")
            print(f"   Workers: 20 threads")
            print(f"{'='*70}\n")

            with ThreadPoolExecutor(max_workers=20) as executor:
                # Submit all tasks
                future_to_domain = {
                    executor.submit(check_single_domain_c99, domain, idx): (domain, idx)
                    for idx, domain in enumerate(domains_to_check, 1)
                }

                # Process completed tasks
                for future in as_completed(future_to_domain):
                    domain, idx = future_to_domain[future]

                    # Update progress
                    search_progress['current'] = idx
                    search_progress['current_domain'] = domain
                    search_progress['message'] = f'Check: {idx}/{len(domains_to_check)} | Có giá trị: {len(found_domains)}/{len(domains_to_check)}'

                    try:
                        result = future.result()
                        if result:
                            # Thread-safe append + SORT NGAY để domains có traffic lên đầu
                            with found_domains_lock:
                                found_domains.append(result)

                                # SORT NGAY theo logic đầy đủ
                                found_domains.sort(key=lambda x: (
                                    not (x.get('organic_traffic', 0) > 0),  # Có traffic lên đầu
                                    # Status priority
                                    {'available': 0, 'pending_delete': 1, 'redemption': 1, 'auction': 2, 'registered': 3}.get(x.get('status_type', 'registered'), 3),
                                    -x.get('organic_traffic', 0),            # Traffic cao hơn
                                    -x.get('domain_rating', 0),              # DR cao hơn
                                    -x.get('backlinks', 0)                   # Backlinks cao hơn
                                ))

                                # GÁN COPY đã sort để frontend hiển thị đúng thứ tự
                                search_progress['domains_found'] = list(found_domains)

                            print(f"[{idx}/{len(domains_to_check)}] ✅ {domain} - DR:{result['domain_rating']}, Traffic:{result['organic_traffic']}")
                        else:
                            print(f"[{idx}/{len(domains_to_check)}] ⏭️  {domain} - Skipped")
                    except Exception as e:
                        print(f"[{idx}/{len(domains_to_check)}] ❌ {domain} - Error: {e}")

            print(f"\n{'='*70}")
            print(f"✅ PARALLEL CHECK COMPLETED!")
            print(f"   Checked: {len(domains_to_check)} domains")
            print(f"   Found: {len(found_domains)} domains with SEO value")
            print(f"{'='*70}\n")

        # Skip keyword mode - C99 only
        pass

        # SORT với LOCK để tránh conflict khi frontend đang đọc
        with found_domains_lock:
            # SORT PRIORITY: CÓ TRAFFIC LÊN TRÊN
            def sort_priority(domain):
                has_traffic = domain.get('organic_traffic', 0) > 0
                status_type = domain.get('status_type', 'registered')

                # Priority: Available > Pending/Redemption > Auction > Registered
                status_priority_map = {
                    'available': 0,          # ✅ Mua ngay
                    'pending_delete': 1,     # ⏳ Backorder
                    'redemption': 1,         # ⚠️ Backorder
                    'auction': 2,            # 🔨 Đấu giá
                    'registered': 3          # 🔒 Không mua được
                }
                status_priority = status_priority_map.get(status_type, 3)

                return (
                    not has_traffic,                 # Có traffic = False (lên trên), không có = True (xuống dưới)
                    status_priority,                 # Status priority
                    -domain['organic_traffic'],      # Traffic cao lên đầu
                    -domain['domain_rating'],        # DR cao lên đầu
                    -domain['backlinks']             # Backlinks cao lên đầu
                )

            found_domains.sort(key=sort_priority)
            # GÁN COPY để frontend đọc dữ liệu đã sort ổn định
            search_progress['domains_found'] = list(found_domains)
        search_progress['status'] = 'completed'

        # Detailed completion message
        if mode == 'c99':
            # Count by status type
            status_counts = {
                'available': len([d for d in found_domains if d.get('status_type') == 'available']),
                'pending_delete': len([d for d in found_domains if d.get('status_type') == 'pending_delete']),
                'redemption': len([d for d in found_domains if d.get('status_type') == 'redemption']),
                'auction': len([d for d in found_domains if d.get('status_type') == 'auction']),
                'registered': len([d for d in found_domains if d.get('status_type') == 'registered'])
            }

            # Count traffic vs no traffic
            with_traffic = len([d for d in found_domains if d.get('organic_traffic', 0) > 0])
            without_traffic = len(found_domains) - with_traffic

            stats_msg = f'''✓ Hoàn thành! Tìm thấy {len(found_domains)} domains có giá trị

📊 THỐNG KÊ (C99 MODE):
- Đã check: {len(domains_to_check)} domains từ C99.NL
- Có giá trị SEO: {len(found_domains)} domains ({len(found_domains)*100//len(domains_to_check) if len(domains_to_check) > 0 else 0}%)
  • Có traffic: {with_traffic} domains
  • Chỉ có DR/Backlinks: {without_traffic} domains
- Không có giá trị: {len(domains_to_check) - len(found_domains)} domains

🎯 PHÂN LOẠI THEO STATUS:
- ✅ Available (mua ngay): {status_counts['available']}
- ⏳ Pending Delete (backorder): {status_counts['pending_delete']}
- ⚠️ Redemption (backorder): {status_counts['redemption']}
- 🔨 Auction (đấu giá): {status_counts['auction']}
- 🔒 Registered (liên hệ): {status_counts['registered']}'''
        else:
            stats_msg = f'''✓ Hoàn thành! Tìm thấy {len(found_domains)} domains

📊 THỐNG KÊ (KEYWORD MODE):
- Đã check: {len(domains_to_check)} domains
- Tìm được: {len(found_domains)} domains ({len(found_domains)*100//len(domains_to_check) if len(domains_to_check) > 0 else 0}%)

❌ BỊ LOẠI:
- Registered: {skipped_registered}
- No history: {skipped_no_history}
- DR < {min_dr}: {skipped_low_dr}'''

        search_progress['message'] = stats_msg

        # Print summary
        print(f"\n{'='*70}")
        print(f"✅ SEARCH COMPLETED - ID: {search_id}")
        print(f"{'='*70}")
        print(f"Checked: {len(domains_to_check)} domains")
        print(f"Found: {len(found_domains)} domains ({len(found_domains)*100//len(domains_to_check) if len(domains_to_check) > 0 else 0}%)")
        print(f"\nSkipped:")
        print(f"  - Registered: {skipped_registered}")
        print(f"  - No history: {skipped_no_history}")
        print(f"  - DR < {min_dr}: {skipped_low_dr}")
        print(f"{'='*70}\n")

        # Export Excel
        if found_domains:
            export_to_excel(found_domains, search_id)
            print(f"✅ Excel exported: results_{search_id}.xlsx\n")

    except Exception as e:
        search_progress['status'] = 'error'
        search_progress['message'] = f'Lỗi: {str(e)}'


def export_to_excel(results: List[Dict], search_id: str):
    """Export kết quả ra Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Expired Domains"

        # Headers
        headers = ["#", "Domain", "DR", "UR", "Traffic", "Backlinks", "Ref Domains",
                   "Snapshots", "First Archive", "Age (Years)", "Price", "Purchase URL"]

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        for idx, domain_data in enumerate(results, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=domain_data['domain'])
            ws.cell(row=idx+1, column=3, value=domain_data['domain_rating'])
            ws.cell(row=idx+1, column=4, value=domain_data['url_rating'])
            ws.cell(row=idx+1, column=5, value=domain_data['organic_traffic'])
            ws.cell(row=idx+1, column=6, value=domain_data['backlinks'])
            ws.cell(row=idx+1, column=7, value=domain_data['referring_domains'])
            ws.cell(row=idx+1, column=8, value=domain_data['snapshot_count'])
            ws.cell(row=idx+1, column=9, value=domain_data['first_archive'])
            ws.cell(row=idx+1, column=10, value=domain_data['age_years'])
            ws.cell(row=idx+1, column=11, value=domain_data['price'])
            ws.cell(row=idx+1, column=12, value=domain_data['purchase_url'])

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save
        filename = f"results_{search_id}.xlsx"
        filepath = os.path.join('static', filename)
        os.makedirs('static', exist_ok=True)
        wb.save(filepath)

        search_progress['excel_file'] = filename

    except Exception as e:
        print(f"Error exporting Excel: {e}")


@app.route('/')
def index():
    """Homepage"""
    return render_template('index.html')


@app.route('/api/suggest-keywords', methods=['POST'])
def suggest_keywords():
    """API suggest related keywords"""
    data = request.json
    keyword = data.get('keyword', '').strip()

    if not keyword:
        return jsonify({'error': 'No keyword provided'}), 400

    # Generate related keywords + variations
    result = keyword_gen.generate_all(keyword, max_total=50)

    # Quick suggestions (top 10)
    suggestions = keyword_gen.suggest_keywords(keyword)

    return jsonify({
        'original': keyword,
        'suggestions': suggestions,
        'related': result['related'],
        'variations': result['variations'][:15],
        'all': result['all']
    })


@app.route('/api/search', methods=['POST'])
def start_search():
    """Bắt đầu search domains"""
    global search_progress

    data = request.json

    mode = data.get('mode', 'keyword')
    c99_domains = data.get('c99_domains', [])
    keywords = [k.strip() for k in data.get('keywords', '').split(',') if k.strip()]
    tlds = data.get('tlds', ['sa.com', 'ru.com', 'in.com', 'za.com', 'br.com'])
    max_check = int(data.get('max_check', 50))
    min_dr = int(data.get('min_dr', 10))

    # Validation based on mode
    if mode == 'c99':
        if not c99_domains:
            return jsonify({'error': 'Vui lòng fetch domains từ C99.NL trước'}), 400
    else:
        if not keywords:
            return jsonify({'error': 'Vui lòng nhập keywords'}), 400

    # Reset progress
    search_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    search_progress = {
        'status': 'starting',
        'current': 0,
        'total': 0,
        'message': 'Đang khởi động...',
        'domains_found': [],
        'current_domain': '',
        'excel_file': None
    }

    # Start background thread
    thread = Thread(target=search_domains_background, args=(keywords, tlds, max_check, min_dr, search_id, mode, c99_domains))
    thread.daemon = True
    thread.start()

    return jsonify({'status': 'started', 'search_id': search_id, 'mode': mode})


@app.route('/api/progress')
def get_progress():
    """Lấy progress hiện tại"""
    return jsonify(search_progress)


@app.route('/api/fetch-c99-domains', methods=['POST'])
def fetch_c99_domains():
    """API fetch domains từ C99.NL"""
    data = request.json
    tld = data.get('tld', '').strip()

    if not tld:
        return jsonify({'error': 'No TLD provided'}), 400

    # Initialize checker
    checker = DomainChecker(rapidapi_key=RAPIDAPI_KEY, c99_api_key=C99_API_KEY)

    # Fetch domains
    domains = checker.fetch_domains_from_c99(tld)

    return jsonify({
        'tld': tld,
        'count': len(domains),
        'domains': domains  # Trả về HẾT tất cả domains (không slice)
    })


@app.route('/api/download/<filename>')
def download_file(filename):
    """Download Excel file"""
    filepath = os.path.join('static', filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


if __name__ == '__main__':
    print("=" * 70)
    print("🌐 DOMAIN FINDER WEB APP")
    print("=" * 70)
    print("\n✓ Server starting...")
    print("✓ Open browser: http://localhost:5000")
    print("\n" + "=" * 70)
    app.run(debug=True, host='0.0.0.0', port=5000)
